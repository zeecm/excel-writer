import os
from tempfile import TemporaryDirectory
from typing import Tuple

import pytest
from openpyxl.styles import Font

from excel_writer.writer import (
    DEFAULT_COLUMN_WIDTH,
    DEFAULT_ROW_HEIGHT,
    CellRange,
    ExcelWriter,
)


class TestExcelWriter:
    def setup_method(self):
        self.writer = ExcelWriter(default_sheet_name="original_sheet")

    def test_writer_class(self):
        assert self.writer.worksheets == ("original_sheet",)

    def test_rename_sheet(self):
        self.writer.rename_sheet(0, new_sheet_name="new_sheet_name")
        assert self.writer.worksheets == ("new_sheet_name",)

    def test_create_sheet(self):
        self.writer.create_sheet("new_sheet")
        assert self.writer.worksheets == ("original_sheet", "new_sheet")

    @pytest.mark.parametrize(
        "sheet_names, expected_order",
        [
            (
                ("sheet_1", "sheet_2", "sheet_3"),
                ("original_sheet", "sheet_1", "sheet_2", "sheet_3"),
            ),
            (
                ("sheet_5", "sheet_4", "sheet_3"),
                ("original_sheet", "sheet_5", "sheet_4", "sheet_3"),
            ),
        ],
    )
    def test_worksheet_property_ordered(
        self, sheet_names: Tuple[str, ...], expected_order: Tuple[str, ...]
    ):
        for sheet in sheet_names:
            self.writer.create_sheet(sheet)
        assert self.writer.worksheets == expected_order

    def test_save_workbook(self):
        with TemporaryDirectory() as tmpdir:
            self.writer.save_workbook(filepath=tmpdir, filename="test.xlsx")
            assert os.path.isfile(os.path.join(tmpdir, "test.xlsx"))

    def test_set_current_sheet_by_name(self):
        self.writer.create_sheet("new_sheet")
        self.writer.set_active_sheet("new_sheet")
        assert self.writer.active_sheet.title == "new_sheet"

    def test_set_current_sheet_by_index(self):
        self.writer.create_sheet("new_sheet_1")
        self.writer.create_sheet("new_sheet_2")
        self.writer.set_active_sheet(2)
        assert self.writer.active_sheet.title == "new_sheet_2"

    def test_cell_using_notation(self):
        cell = self.writer.cell(0, "A1", set_value="test_value")
        assert cell.value == "test_value"

    def test_cell_using_row_col(self):
        cell = self.writer.cell(0, (1, 2), set_value="test_value")
        assert cell.value == "test_value"

    def test_move_range_with_dimensions(self):
        # type ignores for openpyxl row dimension indexing
        self.writer.cell(0, "A1", set_value="a1")
        self.writer.cell(0, "B2", set_value="b2")
        self.writer.active_sheet.row_dimensions[1].height = 30  # type: ignore
        self.writer.active_sheet.column_dimensions["A"].width = 50
        cell_range = CellRange(start_row=1, end_row=2, start_column=1, end_column=2)
        self.writer.move_range(
            0, cell_range, rows_to_move=5, columns_to_move=5, move_dimensions=True
        )
        assert self.writer.cell(0, "F6").value == "a1"
        assert self.writer.cell(0, "G7").value == "b2"
        assert self.writer.active_sheet.row_dimensions[6].height == 30  # type: ignore
        assert self.writer.active_sheet.column_dimensions["F"].width == 50
        assert self.writer.active_sheet.row_dimensions[1].height == DEFAULT_ROW_HEIGHT  # type: ignore
        assert (
            self.writer.active_sheet.column_dimensions["A"].width
            == DEFAULT_COLUMN_WIDTH
        )

    def test_get_cell_style(self):
        ft = Font(color="FF0000")
        self.writer.cell(0, "A1").font = ft
        cell_style = self.writer.cell_style(0, "A1")
        assert cell_style.font == ft


@pytest.mark.parametrize(
    "start_row, start_column, end_row, end_column, expected_notation",
    [
        (1, 1, 2, 2, "A1:B2"),
        (1, 2, 3, 4, "B1:D3"),
        (10, 10, 50, 50, "J10:AX50"),
    ],
)
def test_cell_range_notation(
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
    expected_notation: str,
):
    cell_range = CellRange(
        start_row=start_row,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column,
    )
    assert cell_range.notation == expected_notation
