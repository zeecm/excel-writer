from __future__ import annotations

import os
import traceback
from copy import copy
from shutil import rmtree
from tempfile import mkdtemp
from typing import (
    Any,
    Dict,
    List,
    NamedTuple,
    Optional,
    Protocol,
    Tuple,
    Type,
    Union,
    overload,
)

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_IS_WINDOWS = False

try:
    from win32com import client

    _IS_WINDOWS = True
except ImportError:
    logger.warning("not on windows, no win32com client")

_CellTypes = Type[Cell]

DEFAULT_ROW_HEIGHT = 15
DEFAULT_COLUMN_WIDTH = 8.43


# Cell Row and Column integeres are 1-based indexed


class CellStyle(NamedTuple):
    font: Font
    fill: PatternFill
    border: Border
    alignment: Alignment


class CellRange(NamedTuple):
    start_row: int
    start_column: int
    end_row: int
    end_column: int

    def move_range(self, rows_to_move: int = 0, columns_to_move: int = 0) -> CellRange:
        return CellRange(
            start_row=self.start_row + rows_to_move,
            end_row=self.end_row + rows_to_move,
            start_column=self.start_column + columns_to_move,
            end_column=self.end_column + columns_to_move,
        )

    @property
    def notation(self) -> str:
        return f"{self.start_notation}:{self.end_notation}"

    @property
    def start_notation(self) -> str:
        start_row_letter = get_column_letter(self.start_column)
        return f"{start_row_letter}{self.start_row}"

    @property
    def end_notation(self) -> str:
        end_row_letter = get_column_letter(self.end_column)
        return f"{end_row_letter}{self.end_row}"


class Writer(Protocol):
    worksheets: Tuple[str, ...]

    def cell(
        self,
        sheet: Union[str, int],
        cell_id: Union[Tuple[int, int], str],
        set_value: Optional[str] = None,
    ) -> _CellTypes:
        ...

    def save_workbook(self, filepath: str, filename: str) -> None:
        ...


class ExcelWriter:
    def __init__(
        self,
        existing_workbook: str = "",
        default_sheet_name: str = "",
        iso_dates: bool = False,
        default_row_height: float = DEFAULT_ROW_HEIGHT,
        default_column_width: float = DEFAULT_COLUMN_WIDTH,
    ):
        self._workbook = self._initialize_workbook(
            existing_workbook, default_sheet_name, iso_dates
        )
        self.active_sheet = self._get_active_sheet(self._workbook)
        self._default_row_height = default_row_height
        self._default_column_width = default_column_width

    def _initialize_workbook(
        self,
        existing_workbook: str = "",
        default_sheet_name: str = "",
        iso_dates: bool = False,
    ) -> Workbook:
        if existing_workbook:
            return self._load_existing_workbook(existing_workbook)
        return self._create_new_workbook(default_sheet_name, iso_dates)

    def _load_existing_workbook(self, filepath: str) -> Workbook:
        return load_workbook(filepath)

    def _create_new_workbook(
        self, default_sheet_name: str = "", iso_dates: bool = False
    ) -> Workbook:
        workbook = Workbook(iso_dates=iso_dates)
        active_sheet = self._get_active_sheet(workbook)
        if default_sheet_name:
            active_sheet.title = default_sheet_name
        return workbook

    def _get_active_sheet(self, workbook: Workbook) -> Worksheet:
        if not workbook.active:
            workbook.active = self.get_worksheet(0)
        # correct typing, worksheet is subclass of _WorkbookChild
        return workbook.active  # type: ignore

    @property
    def worksheets(self) -> Tuple[str, ...]:
        return tuple(self._workbook.sheetnames)

    def create_sheet(self, sheet_name: str, position: Optional[int] = None) -> None:
        self._workbook.create_sheet(sheet_name, position)
        self.set_active_sheet(sheet_name)

    def rename_sheet(self, sheet: Union[str, int], new_sheet_name: str) -> None:
        sheet_obj = self.get_worksheet(sheet)
        sheet_obj.title = new_sheet_name

    def save_workbook(self, filepath: str, filename: str) -> None:
        full_filepath = os.path.join(filepath, filename)
        self._workbook.save(full_filepath)
        logger.info(f"saved workbook to {full_filepath}")

    def export_as_pdf(
        self, filepath: str, filename: str, sheet: Union[str, int] = 0
    ) -> None:
        pdf_filepath = os.path.join(filepath, filename)
        sheet_index = self._get_sheet_index(sheet)
        if not _IS_WINDOWS:
            logger.error("Unable to export as PDF, not on Windows")
            return
        self._save_temporary_excel_and_print_pdf(sheet_index, pdf_filepath)

    def _save_temporary_excel_and_print_pdf(
        self, sheet_index: int, pdf_filepath: str
    ) -> None:
        temp_excel_filename = self._generate_temp_workbook_filename(pdf_filepath)
        tmpdir = mkdtemp()
        temp_filepath = os.path.join(tmpdir, temp_excel_filename)
        self.save_workbook(tmpdir, temp_excel_filename)
        excel = client.Dispatch("Excel.Application")
        self._print_pdf_with_error_handling(
            excel, temp_filepath, pdf_filepath, sheet_index
        )
        rmtree(tmpdir, ignore_errors=True)
        excel.Quit()

    def _generate_temp_workbook_filename(self, pdf_filepath: str) -> str:
        filename_only = os.path.basename(pdf_filepath)
        return filename_only.replace(".pdf", ".xlsx")

    def _get_sheet_index(self, sheet: Union[str, int]) -> int:
        worksheet = self.get_worksheet(sheet)
        return self._workbook.index(worksheet)

    def _print_pdf_with_error_handling(
        self,
        excel_client: Any,
        temp_workbook_filepath: str,
        pdf_filepath: str,
        sheet_index: int,
    ) -> None:
        try:
            self._print_pdf_using_win32com_client(
                excel_client, temp_workbook_filepath, pdf_filepath, sheet_index
            )
        except AttributeError:
            logger.error(f"failed to save to PDF: {traceback.format_exc()}")

    def _print_pdf_using_win32com_client(
        self,
        excel_client: Any,
        temp_workbook_filepath: str,
        pdf_filepath: str,
        sheet_index: int,
    ) -> None:
        excel_client.Visible = False
        workbook = excel_client.Workbooks.Open(temp_workbook_filepath)
        worksheet = workbook.Worksheets[sheet_index]
        worksheet.ExportAsFixedFormat(0, pdf_filepath)
        logger.info(f"saved pdf to {pdf_filepath}")
        workbook.Saved = True
        workbook.Close()

    @overload
    def cell(
        self,
        sheet: Union[str, int],
        cell_id: Tuple[int, int],
        set_value: Optional[str] = None,
        set_style: Optional[CellStyle] = None,
    ) -> Cell:
        ...

    @overload
    def cell(
        self,
        sheet: Union[str, int],
        cell_id: str,
        set_value: Optional[str] = None,
        set_style: Optional[CellStyle] = None,
    ) -> Cell:
        ...

    def cell(
        self,
        sheet: Union[str, int],
        cell_id: Union[Tuple[int, int], str],
        set_value: Optional[str] = None,
        set_style: Optional[CellStyle] = None,
    ) -> Cell:
        sheet_object = self.get_worksheet(sheet)

        if isinstance(cell_id, tuple):
            cell = self._get_cell_by_row_col(sheet_object, row_col=cell_id)
        elif isinstance(cell_id, str):
            cell = self._get_cell_by_notation(sheet_object, cell_notation=cell_id)
        else:
            raise ValueError("one of row and column or cell notation must be specified")
        if set_value:
            cell.value = set_value
        if set_style:
            cell = self._set_cell_style(cell, set_style)
        return cell

    def _get_cell_by_row_col(self, sheet: Worksheet, row_col: Tuple[int, int]) -> Cell:
        row, column = row_col
        return sheet.cell(row=row, column=column)

    def _get_cell_by_notation(self, sheet: Worksheet, cell_notation: str) -> Cell:
        return sheet[cell_notation]

    def _set_cell_style(self, cell: Cell, style: CellStyle) -> Cell:
        cell.font = style.font
        cell.fill = style.fill
        cell.border = style.border
        cell.alignment = style.alignment

        return cell

    @overload
    def get_worksheet(self, sheet: str) -> Worksheet:
        ...

    @overload
    def get_worksheet(self, sheet: int) -> Worksheet:
        ...

    def get_worksheet(self, sheet: Union[str, int]) -> Worksheet:
        if isinstance(sheet, str):
            return self._workbook[sheet]
        if isinstance(sheet, int):
            return self._workbook.worksheets[sheet]
        raise ValueError(f"invalid sheet argument {sheet}")

    def set_active_sheet(self, sheet: Optional[Union[str, int]] = None) -> None:
        """Sets current sheet, if sheet is not provided, defaults to first sheet"""
        if sheet is None:
            self.active_sheet = self._workbook.worksheets[0]
            return
        self.active_sheet = self.get_worksheet(sheet)

    def cell_style(
        self,
        sheet: Union[str, int],
        cell_id: Union[Tuple[int, int], str],
    ) -> CellStyle:
        cell = self.cell(sheet, cell_id)

        font = copy(cell.font)
        border = copy(cell.border)
        fill = copy(cell.fill)
        alignment = copy(cell.alignment)

        return CellStyle(font=font, fill=fill, border=border, alignment=alignment)

    def move_range(
        self,
        sheet: Union[str, int],
        cell_range: CellRange,
        rows_to_move: int = 0,
        columns_to_move: int = 0,
        translate_formulas: bool = False,
        move_dimensions: bool = True,
    ) -> CellRange:
        worksheet = self.get_worksheet(sheet)
        range_notation = cell_range.notation
        if move_dimensions:
            self._move_range_dimensions_to_new_range(
                sheet, cell_range, rows_to_move, columns_to_move
            )
        worksheet.move_range(
            range_notation,
            rows=rows_to_move,
            cols=columns_to_move,
            translate=translate_formulas,
        )
        return cell_range.move_range(
            rows_to_move=rows_to_move, columns_to_move=columns_to_move
        )

    def _move_range_dimensions_to_new_range(
        self,
        sheet: Union[str, int],
        cell_range: CellRange,
        rows_to_move: int = 0,
        columns_to_move: int = 0,
    ) -> None:
        logger.debug("moving range with styles")
        row_height_map = self._copy_range_row_height(sheet, cell_range)
        column_width_map = self._copy_range_column_width(sheet, cell_range)
        self._set_old_row_height_to_default(sheet, cell_range)
        self._set_old_column_width_to_default(sheet, cell_range)
        self._set_range_row_height(sheet, row_height_map, rows_moved=rows_to_move)
        self._set_range_column_width(
            sheet, column_width_map, columns_moved=columns_to_move
        )

    def _copy_range_row_height(
        self,
        sheet: Union[str, int],
        cell_range: CellRange,
    ) -> Dict[int, float]:
        worksheet = self.get_worksheet(sheet)
        current_row = cell_range.start_row
        end_row = cell_range.end_row

        row_height_map = {}

        while current_row <= end_row:
            # openpyxl row dimension indexing
            row_dimensions = worksheet.row_dimensions[current_row].height  # type: ignore
            row_height_map[current_row] = row_dimensions
            current_row += 1
        return row_height_map

    def _copy_range_column_width(
        self,
        sheet: Union[str, int],
        cell_range: CellRange,
    ) -> Dict[int, float]:
        worksheet = self.get_worksheet(sheet)
        current_column = cell_range.start_column
        end_column = cell_range.end_column

        column_width_map = {}

        while current_column <= end_column:
            column_letter = get_column_letter(current_column)
            column_width = worksheet.column_dimensions[column_letter].width
            column_width_map[current_column] = column_width
            current_column += 1
        return column_width_map

    def _set_old_row_height_to_default(
        self,
        sheet: Union[str, int],
        cell_range: CellRange,
        default_row_height: Optional[float] = None,
    ) -> None:
        default_row_height = default_row_height or self._default_row_height
        worksheet = self.get_worksheet(sheet)
        current_row = cell_range.start_row
        end_row = cell_range.end_row

        while current_row <= end_row:
            # openpyxl row dimension indexing
            worksheet.row_dimensions[current_row].height = default_row_height  # type: ignore
            current_row += 1

    def _set_old_column_width_to_default(
        self,
        sheet: Union[str, int],
        cell_range: CellRange,
        default_column_width: Optional[float] = None,
    ) -> None:
        default_column_width = default_column_width or self._default_column_width
        worksheet = self.get_worksheet(sheet)
        current_column = cell_range.start_column
        end_column = cell_range.end_column

        while current_column <= end_column:
            column_letter = get_column_letter(current_column)
            worksheet.column_dimensions[column_letter].width = default_column_width
            current_column += 1

    def _set_range_row_height(
        self,
        sheet: Union[str, int],
        row_height_map: Dict[int, float],
        rows_moved: int = 0,
    ) -> None:
        worksheet = self.get_worksheet(sheet)
        for row, height in row_height_map.items():
            new_row = row + rows_moved
            # openpyxl row dimension indexing
            worksheet.row_dimensions[new_row].height = height  # type: ignore

    def _set_range_column_width(
        self,
        sheet: Union[str, int],
        column_width_map: Dict[int, float],
        columns_moved: int = 0,
    ) -> None:
        worksheet = self.get_worksheet(sheet)
        for column, width in column_width_map.items():
            new_column = column + columns_moved
            column_letter = get_column_letter(new_column)
            worksheet.column_dimensions[column_letter].width = width

    def get_print_area(
        self,
        sheet: Union[str, int] = 0,
    ) -> List[str]:
        worksheet = self.get_worksheet(sheet)
        return worksheet.print_area

    def set_print_area(
        self,
        sheet: Union[str, int],
        print_area: Union[str, List[str]],
    ) -> None:
        worksheet = self.get_worksheet(sheet)
        worksheet.print_area = print_area
